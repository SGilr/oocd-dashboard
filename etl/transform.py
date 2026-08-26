#!/usr/bin/env python3
"""Read the raw Home Office files and write the derived tables the site uses.

Nothing here is positional. Headers are introspected and normalised, because
they vary between years, and a header that cannot be mapped stops the run
rather than being dropped in silence.

Both count bases are carried through every derived table. Percentages are not
computed here, they are computed at render time, so the stored numbers stay
integers and the rounding is visible in one place.

Usage:
    python etl/transform.py
    python etl/transform.py --raw-dir data/raw --out-dir data/processed
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import (  # noqa: E402
    NOT_ASSIGNED_TYPE,
    BASIS_CLOSED,
    BASIS_RECORDED,
    CENTRAL_FRAUD_BODIES,
    COUNT_BASES,
    DEFAULT_BASIS,
    OOCD_TYPES,
    OUTCOME_LABELS,
    POSITIVE_TYPES,
    REQUIRED_OUTCOME_FIELDS,
    HeaderMappingError,
    build_header_map,
    canonical_force,
    financial_year_start,
    is_not_provided,
    force_slug,
    normalise_financial_year,
    normalise_quarter,
    read_json,
    resolve_data_root,
    text_of,
    truthy_flag,
    to_int,
    truthy_flag,
    write_json_compact,
)

# The outcome types stored individually in the derived tables: the charge
# comparator and the six out of court types. Everything else is carried only
# inside the assigned total, which is what the denominators need.
STORED_TYPES: tuple[int, ...] = POSITIVE_TYPES
COUNT_COLUMNS: tuple[str, ...] = tuple(f"t{t}" for t in STORED_TYPES) + (
    "oocd",
    "positive",
    "assigned",
)

FRAUD_VARIANTS: tuple[str, ...] = ("all", "ex_fraud")


def is_fraud_group(offence_group: str | None, offence_subgroup: str | None) -> bool:
    """True when a row belongs to the fraud offence group.

    Fraud reported to Action Fraud, Cifas and Financial Fraud UK is recorded
    centrally and then attributed to force areas, so it distorts offence mix.
    The exclude fraud variant drops the whole fraud group rather than trying to
    separate centrally recorded fraud from force recorded fraud, which the
    published tables do not support.
    """
    for value in (offence_group, offence_subgroup):
        if value and "fraud" in str(value).lower():
            return True
    return False


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------


def read_csv_rows(path: Path) -> Iterator[list[object]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.reader(handle):
            yield row


def read_xlsx_rows(path: Path) -> Iterator[list[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _pick_sheet(workbook)
        for row in sheet.iter_rows(values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _pick_sheet(workbook):
    """Choose the data sheet, skipping cover, notes and contents sheets."""
    for sheet in workbook.worksheets:
        if _skippable_sheet(sheet.title):
            continue
        return sheet
    return workbook.worksheets[0]


# OpenDocument namespaces, needed to read the police force area crime tables,
# which are published as ODS rather than as xlsx.
ODS_TABLE = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"
ODS_OFFICE = "{urn:oasis:names:tc:opendocument:xmlns:office:1.0}"
ODS_TEXT = "{urn:oasis:names:tc:opendocument:xmlns:text:1.0}"

# A repeat count above this is the spreadsheet padding out to its full grid
# rather than real data. Expanding those would produce millions of empty cells.
ODS_PADDING_REPEAT = 512


def _ods_cell_value(cell) -> object:
    """Read one ODS cell, preferring the typed value over the displayed text."""
    value_type = cell.get(f"{ODS_OFFICE}value-type")
    if value_type in {"float", "percentage", "currency"}:
        raw = cell.get(f"{ODS_OFFICE}value")
        if raw is not None:
            number = float(raw)
            return int(number) if number.is_integer() else number
    if value_type == "boolean":
        return cell.get(f"{ODS_OFFICE}boolean-value")
    if value_type == "date":
        return cell.get(f"{ODS_OFFICE}date-value")
    text = "".join(node.text or "" for node in cell.iter(f"{ODS_TEXT}p"))
    return text or None


def read_ods_sheets(path: Path) -> Iterator[tuple[str, Iterator[list[object]]]]:
    """Yield (sheet name, rows) for every sheet in an ODS file.

    Streamed with iterparse straight out of the zip, because these files are
    tens of megabytes compressed and expand to far more as XML. Nothing is held
    beyond the row being built.
    """
    import zipfile
    from xml.etree.ElementTree import iterparse

    with zipfile.ZipFile(path) as archive:
        with archive.open("content.xml") as content:
            sheet_name = ""
            rows: list[list[object]] = []
            for event, element in iterparse(content, events=("start", "end")):
                if event == "start" and element.tag == f"{ODS_TABLE}table":
                    sheet_name = element.get(f"{ODS_TABLE}name", "")
                    rows = []
                elif event == "end" and element.tag == f"{ODS_TABLE}table-row":
                    cells: list[object] = []
                    for cell in element.findall(f"{ODS_TABLE}table-cell"):
                        value = _ods_cell_value(cell)
                        repeat = int(cell.get(f"{ODS_TABLE}number-columns-repeated", 1))
                        if value is None and repeat > ODS_PADDING_REPEAT:
                            repeat = 1
                        cells.extend([value] * repeat)
                    while cells and cells[-1] is None:
                        cells.pop()
                    row_repeat = int(
                        element.get(f"{ODS_TABLE}number-rows-repeated", 1)
                    )
                    if not cells and row_repeat > ODS_PADDING_REPEAT:
                        row_repeat = 1
                    for _ in range(row_repeat):
                        rows.append(list(cells))
                    element.clear()
                elif event == "end" and element.tag == f"{ODS_TABLE}table":
                    yield sheet_name, iter(rows)
                    rows = []
                    element.clear()


def read_ods_rows(path: Path) -> Iterator[list[object]]:
    """Rows of the first ODS sheet that is not a cover or notes sheet."""
    for sheet_name, rows in read_ods_sheets(path):
        if _skippable_sheet(sheet_name):
            continue
        yield from rows
        return


SKIP_SHEET_TOKENS = ("cover", "note", "contents", "metadata", "definitions", "guide")


def _skippable_sheet(name: str) -> bool:
    lowered = str(name).strip().lower()
    return any(token in lowered for token in SKIP_SHEET_TOKENS)


def read_rows(path: Path) -> Iterator[list[object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return read_xlsx_rows(path)
    if suffix == ".ods":
        return read_ods_rows(path)
    raise ValueError(f"Cannot read {path.name}: unsupported file type {suffix}")


def find_header_row(rows: Iterator[list[object]], required: tuple[str, ...]):
    """Scan the first rows for the header, so a cover block does not break it.

    Consumes rows from the iterator up to and including the header row, and
    returns (header_map, header_row_values). The caller keeps iterating the
    same iterator to read the data. Raises when no row in the first twenty
    looks like a header.
    """
    seen = 0
    last_error: Exception | None = None
    for row in rows:
        seen += 1
        if seen > 20:
            break
        try:
            header_map = build_header_map(row, required=required)
        except HeaderMappingError as error:
            last_error = error
            continue
        return header_map, row
    raise HeaderMappingError(
        "No header row found in the first 20 rows. Last attempt said: "
        f"{last_error}"
    )


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------


class Aggregator:
    """Accumulates counts while streaming the raw rows."""

    def __init__(self) -> None:
        # (force, fy, quarter) -> basis -> variant -> {column: count}
        self.quarter: dict[tuple[str, str, int], dict] = {}
        # (force, fy, offence_group) -> basis -> {oocd, charge}
        self.offence: dict[tuple[str, str, str], dict] = {}
        self.outcome_type_years: dict[int, set[str]] = defaultdict(set)
        self.unknown_forces: dict[str, int] = defaultdict(int)
        self.central_fraud_rows = 0
        self.expired_code_rows = 0
        self.rows_read = 0
        self.rows_used = 0
        self.negative_counts = 0
        self.duplicate_source_keys = 0
        self.duplicate_examples: list[str] = []
        self.negative_by_key: dict[tuple[str, str], int] = defaultdict(int)
        self.negative_excluded_examples: list[str] = []
        self.not_assigned_rows = 0
        self.not_assigned_counts = {BASIS_CLOSED: 0, BASIS_RECORDED: 0}
        self.not_provided: dict[tuple[str, str], int] = defaultdict(int)
        self.negative_counts_excluded = 0
        self.negative_examples: list[str] = []
        self.subset_type_rows: dict[str, int] = defaultdict(int)

    @staticmethod
    def _empty_cell() -> dict:
        return {
            basis: {
                variant: dict.fromkeys(COUNT_COLUMNS, 0) for variant in FRAUD_VARIANTS
            }
            for basis in COUNT_BASES
        }

    def add(
        self,
        force: str,
        financial_year: str,
        quarter: int,
        offence_group: str,
        outcome_type: int,
        counts: dict[str, int],
        fraud: bool,
    ) -> None:
        cell = self.quarter.setdefault(
            (force, financial_year, quarter), self._empty_cell()
        )
        variants = ("all",) if fraud else FRAUD_VARIANTS

        for basis in COUNT_BASES:
            value = counts[basis]
            if value == 0:
                continue
            for variant in variants:
                bucket = cell[basis][variant]
                bucket["assigned"] += value
                if outcome_type in STORED_TYPES:
                    bucket[f"t{outcome_type}"] += value
                    bucket["positive"] += value
                    if outcome_type in OOCD_TYPES:
                        bucket["oocd"] += value

        if outcome_type in POSITIVE_TYPES:
            offence_cell = self.offence.setdefault(
                (force, financial_year, offence_group),
                {basis: {"oocd": 0, "charge": 0} for basis in COUNT_BASES},
            )
            for basis in COUNT_BASES:
                value = counts[basis]
                if value == 0:
                    continue
                key = "oocd" if outcome_type in OOCD_TYPES else "charge"
                offence_cell[basis][key] += value


def _note_negative(aggregator, path, force, financial_year, quarter, offence_code,
                   outcome_type, counts, excluded: bool = False) -> None:
    """Record a negative count so a person can see exactly where it was.

    Kept in two lists. Negatives inside outcome type 0 never reach a derived
    total and are far more numerous, so mixing them with the ones that do would
    fill the sample and hide the rows that matter.
    """
    entry = (
        f"{path.name}: {force}, {financial_year} Q{quarter}, offence "
        f"{offence_code}, outcome type {outcome_type}, "
        f"recorded {counts[BASIS_RECORDED]}, closed {counts[BASIS_CLOSED]}"
    )
    target = (
        aggregator.negative_excluded_examples if excluded else aggregator.negative_examples
    )
    if len(target) < 25:
        target.append(entry)
    if not excluded:
        aggregator.negative_by_key[(force, financial_year)] += 1


def process_outcomes_file(path: Path, aggregator: Aggregator, log: list[str]) -> None:
    """Stream one outcomes file into the aggregator.

    The workbook is searched for its data sheet rather than the first sheet
    being assumed. The earlier years carry summary and chart sheets alongside
    the data, and a summary sheet's first rows are figures, not headers, so
    picking by position fails. A sheet counts as data when its header row maps
    every field the ETL needs, which no summary sheet does.
    """

    def emit(message: str) -> None:
        log.append(message)
        print(message, flush=True)

    emit(f"Reading {path.name}")

    tried: list[str] = []
    data_sheets = 0

    for sheet_name, rows in _sheets_of(path):
        if _skippable_sheet(sheet_name):
            emit(f"  sheet {sheet_name!r}: skipped, not a data sheet")
            continue
        try:
            header_map, header_row = find_header_row(rows, REQUIRED_OUTCOME_FIELDS)
        except HeaderMappingError as error:
            tried.append(f"{sheet_name}: {error}")
            emit(f"  sheet {sheet_name!r}: no usable header, skipped")
            continue

        data_sheets += 1
        emit(f"  sheet {sheet_name!r}: headers mapped, reading")
        _read_outcome_rows(path, sheet_name, rows, header_row, header_map, aggregator)

    if data_sheets == 0:
        raise HeaderMappingError(
            f"{path.name} has no sheet whose headers map to the fields the ETL "
            "needs. Sheets tried:\n  " + "\n  ".join(tried)
        )


def _read_outcome_rows(
    path: Path,
    sheet_name: str,
    rows: Iterator[list[object]],
    header_row: list[object],
    header_map,
    aggregator: Aggregator,
) -> None:
    """Read the rows of one data sheet, the header having already been found."""
    index = {
        field: header_row.index(column) for field, column in header_map.by_field.items()
    }

    def cell(row: list[object], field: str) -> object:
        position = index.get(field)
        if position is None or position >= len(row):
            return None
        return row[position]

    # Duplicate detection is per sheet, because a key cannot legitimately repeat
    # inside one. The key is packed into an integer so a multi million row sheet
    # stays in memory.
    seen_keys: set[int] = set()
    force_index: dict[str, int] = {}
    offence_index: dict[str, int] = {}

    for row in rows:
        if row is None or not any(value not in (None, "") for value in row):
            continue
        aggregator.rows_read += 1

        financial_year = normalise_financial_year(cell(row, "financial_year"))
        quarter = normalise_quarter(cell(row, "financial_quarter"))
        if financial_year is None or quarter is None:
            continue

        raw_force = cell(row, "force_name")
        force = canonical_force(raw_force)
        if force is None:
            aggregator.unknown_forces[text_of(raw_force)] += 1
            continue
        if force in CENTRAL_FRAUD_BODIES:
            aggregator.central_fraud_rows += 1
            continue

        raw_outcome_type = text_of(cell(row, "outcome_type"))
        try:
            outcome_type = int(raw_outcome_type)
        except (TypeError, ValueError):
            # Outcome types 1a, 2a and 3a are "of which" rows: the subset of
            # outcomes 1, 2 and 3 that relate to an alternative offence to the
            # one recorded. They are already inside their parent type, so adding
            # them would double count. They are dropped, and counted here so the
            # drop is visible in coverage.json rather than silent.
            if raw_outcome_type:
                aggregator.subset_type_rows[raw_outcome_type] += 1
            continue

        raw_recorded = cell(row, "count_recorded")
        raw_closed = cell(row, "count_closed")
        if is_not_provided(raw_recorded) or is_not_provided(raw_closed):
            # The force did not supply this count. It reads as zero so the
            # arithmetic works, but the force year is understated and that has
            # to be visible, not absorbed.
            aggregator.not_provided[(force, financial_year)] += 1

        counts = {
            BASIS_RECORDED: to_int(raw_recorded),
            BASIS_CLOSED: to_int(raw_closed),
        }

        if truthy_flag(cell(row, "offence_code_expired")):
            aggregator.expired_code_rows += 1

        offence_code = text_of(cell(row, "offence_code"))
        packed = (
            force_index.setdefault(force, len(force_index)) << 40
            | offence_index.setdefault(offence_code, len(offence_index)) << 16
            | (financial_year_start(financial_year) % 100) << 8
            | (quarter - 1) << 6
            | (outcome_type & 63)
        )
        if packed in seen_keys:
            aggregator.duplicate_source_keys += 1
            if len(aggregator.duplicate_examples) < 10:
                aggregator.duplicate_examples.append(
                    f"{path.name} [{sheet_name}]: {force}, {financial_year} "
                    f"Q{quarter}, offence {offence_code}, outcome {outcome_type}"
                )
        else:
            seen_keys.add(packed)

        offence_group = text_of(cell(row, "offence_group"), "Unclassified")
        offence_subgroup = cell(row, "offence_subgroup")
        fraud = is_fraud_group(offence_group, offence_subgroup)

        if any(counts.values()):
            # A row exists for every outcome type in every year, because the
            # published tables are a dense cross product. Recording the year
            # only when a count is present makes "first appears in" mean
            # something: outcome 22 has rows from the start and counts only
            # from the year it was introduced.
            aggregator.outcome_type_years[outcome_type].add(financial_year)

        if outcome_type == NOT_ASSIGNED_TYPE:
            # Outcome type 0 counts offences that have not been given an outcome
            # yet. It is not an outcome, so it cannot go into all assigned
            # outcomes: doing so would put undecided cases into the denominator
            # of every share measure. It is counted here so the volume is
            # visible in coverage.json rather than simply disappearing.
            aggregator.not_assigned_rows += 1
            for basis in COUNT_BASES:
                aggregator.not_assigned_counts[basis] += counts[basis]
            if any(value < 0 for value in counts.values()):
                aggregator.negative_counts_excluded += 1
                _note_negative(aggregator, path, force, financial_year, quarter,
                               offence_code, outcome_type, counts, excluded=True)
            aggregator.rows_used += 1
            continue

        if any(value < 0 for value in counts.values()):
            aggregator.negative_counts += 1
            _note_negative(aggregator, path, force, financial_year, quarter,
                           offence_code, outcome_type, counts)

        aggregator.add(
            force=force,
            financial_year=financial_year,
            quarter=quarter,
            offence_group=offence_group,
            outcome_type=outcome_type,
            counts=counts,
            fraud=fraud,
        )
        aggregator.rows_used += 1


# ---------------------------------------------------------------------------
# Denominators
# ---------------------------------------------------------------------------


def _sheets_of(path: Path) -> Iterator[tuple[str, Iterator[list[object]]]]:
    """Every sheet in a file, as (name, rows).

    The police force area crime tables are published as ODS and may split the
    series across sheets, so each sheet gets its own header detection rather
    than the first sheet's headers being assumed to hold for the rest.
    """
    if path.suffix.lower() == ".ods":
        yield from read_ods_sheets(path)
        return
    if path.suffix.lower() == ".csv":
        yield path.stem, read_csv_rows(path)
        return
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            yield sheet.title, _rows_of_sheet(sheet)
    finally:
        workbook.close()


def _rows_of_sheet(sheet) -> Iterator[list[object]]:
    """Rows of one worksheet, with the sheet bound as an argument.

    Binding matters: a generator expression closing over the loop variable would
    read whichever sheet the loop had reached by the time it was first iterated.
    """
    for row in sheet.iter_rows(values_only=True):
        yield list(row)


def process_force_area_crime(paths: list[Path], log: list[str]) -> dict:
    """Read recorded crime totals per force per financial year.

    Population is deliberately absent. No published mid year population
    estimate per police force area could be retrieved and verified when these
    tables were built, so the per capita measure is omitted rather than
    estimated. See docs/METHODOLOGY.md.
    """
    totals: dict[tuple[str, str], int] = defaultdict(int)
    sheets_read = 0
    sheets_skipped = 0

    def emit(message: str) -> None:
        log.append(message)
        print(message, flush=True)

    for path in paths:
        emit(f"Reading {path.name} for the recorded crime denominator")
        for sheet_name, rows in _sheets_of(path):
            if _skippable_sheet(sheet_name):
                emit(f"  sheet {sheet_name!r}: skipped, not a data sheet")
                sheets_skipped += 1
                continue
            try:
                header_map, header_row = find_header_row(
                    rows,
                    required=("financial_year", "force_name", "recorded_crime_count"),
                )
            except HeaderMappingError as error:
                emit(f"  sheet {sheet_name!r}: skipped, no header found")
                sheets_skipped += 1
                continue

            index = {
                field: header_row.index(column)
                for field, column in header_map.by_field.items()
            }

            def cell(row: list[object], field: str) -> object:
                position = index.get(field)
                if position is None or position >= len(row):
                    return None
                return row[position]

            used = 0
            for row in rows:
                if row is None or not any(value not in (None, "") for value in row):
                    continue
                financial_year = normalise_financial_year(cell(row, "financial_year"))
                force = canonical_force(cell(row, "force_name"))
                if financial_year is None or force is None:
                    continue
                if force in CENTRAL_FRAUD_BODIES:
                    continue
                totals[(force, financial_year)] += to_int(
                    cell(row, "recorded_crime_count")
                )
                used += 1
            emit(f"  sheet {sheet_name!r}: {used} rows used")
            sheets_read += 1

    if paths and not totals:
        emit(
            "  WARNING: no recorded crime totals were read. The rate per 1,000 "
            "recorded crimes measure will have no denominator."
        )

    return {
        "meta": {
            "recorded_crime_source": "Home Office police force area crime tables",
            "sheets_read": sheets_read,
            "sheets_skipped": sheets_skipped,
            "population_source": None,
            "population_omitted_because": (
                "No mid year population estimate per police force area could be "
                "retrieved from a named published source when these tables were "
                "built. Per capita measures are omitted rather than estimated."
            ),
        },
        "rows": [
            {
                "force": force,
                "slug": force_slug(force),
                "fy": financial_year,
                "recorded_crime": count,
            }
            for (force, financial_year), count in sorted(totals.items())
        ],
    }


# ---------------------------------------------------------------------------
# Output shaping
# ---------------------------------------------------------------------------


def _sum_cells(cells: list[dict]) -> dict:
    total = Aggregator._empty_cell()
    for cell in cells:
        for basis in COUNT_BASES:
            for variant in FRAUD_VARIANTS:
                for column in COUNT_COLUMNS:
                    total[basis][variant][column] += cell[basis][variant][column]
    return total


def _flatten(cell: dict) -> dict:
    """Flatten a nested count cell into basis_variant_column keys."""
    flat: dict[str, int] = {}
    for basis in COUNT_BASES:
        for variant in FRAUD_VARIANTS:
            for column in COUNT_COLUMNS:
                flat[f"{basis}_{variant}_{column}"] = cell[basis][variant][column]
    return flat


FLAT_COLUMNS: tuple[str, ...] = tuple(
    f"{basis}_{variant}_{column}"
    for basis in COUNT_BASES
    for variant in FRAUD_VARIANTS
    for column in COUNT_COLUMNS
)


def build_tables(aggregator: Aggregator) -> dict[str, dict]:
    force_quarter_rows = []
    year_cells: dict[tuple[str, str], list[dict]] = defaultdict(list)
    national_quarter_cells: dict[tuple[str, int], list[dict]] = defaultdict(list)

    for (force, financial_year, quarter), cell in sorted(aggregator.quarter.items()):
        force_quarter_rows.append(
            {
                "force": force,
                "slug": force_slug(force),
                "fy": financial_year,
                "q": quarter,
                **_flatten(cell),
            }
        )
        year_cells[(force, financial_year)].append(cell)
        national_quarter_cells[(financial_year, quarter)].append(cell)

    force_year_rows = [
        {
            "force": force,
            "slug": force_slug(force),
            "fy": financial_year,
            **_flatten(_sum_cells(cells)),
        }
        for (force, financial_year), cells in sorted(year_cells.items())
    ]

    national_quarter_rows = [
        {"fy": financial_year, "q": quarter, **_flatten(_sum_cells(cells))}
        for (financial_year, quarter), cells in sorted(national_quarter_cells.items())
    ]

    national_year_cells: dict[str, list[dict]] = defaultdict(list)
    for (force, financial_year), cells in year_cells.items():
        national_year_cells[financial_year].extend(cells)
    national_year_rows = [
        {"fy": financial_year, **_flatten(_sum_cells(cells))}
        for financial_year, cells in sorted(national_year_cells.items())
    ]

    force_offence_year_rows = [
        {
            "force": force,
            "slug": force_slug(force),
            "fy": financial_year,
            "offence_group": offence_group,
            **{
                f"{basis}_{key}": cell[basis][key]
                for basis in COUNT_BASES
                for key in ("oocd", "charge")
            },
        }
        for (force, financial_year, offence_group), cell in sorted(
            aggregator.offence.items()
        )
    ]

    return {
        "force_quarter": force_quarter_rows,
        "force_year": force_year_rows,
        "national_quarter": national_quarter_rows,
        "national_year": national_year_rows,
        "force_offence_year": force_offence_year_rows,
    }


def write_table(out_dir: Path, name: str, rows: list[dict], meta: dict) -> None:
    payload = {"meta": meta, "rows": rows}
    write_json_compact(out_dir / f"{name}.json", payload)
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    with (out_dir / f"{name}.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def survey(paths, log: list[str]) -> dict:
    """Report every surprise in the raw files without building anything.

    Twelve years of published files do not use one convention. Running the full
    transform and fixing each crash in turn costs a ten minute run per fix, so
    this reads everything once and reports every distinct value the ETL would
    have to interpret: the not applicable markers, the expired flags, the
    outcome type values, the force names that do not match, and the sheets in
    each workbook. One pass, then fix everything at once.
    """
    from collections import Counter, defaultdict

    findings = {
        "sheets": defaultdict(list),
        "non_numeric_counts": Counter(),
        "expired_flags": Counter(),
        "outcome_types": Counter(),
        "unmatched_forces": Counter(),
        "financial_years": Counter(),
        "quarters": Counter(),
        "headers": defaultdict(list),
    }

    for path in sorted(paths):
        print(f"Surveying {path.name}", flush=True)
        for sheet_name, rows in _sheets_of(path):
            if _skippable_sheet(sheet_name):
                findings["sheets"][path.name].append(f"{sheet_name} (skipped)")
                continue
            try:
                header_map, header_row = find_header_row(rows, REQUIRED_OUTCOME_FIELDS)
            except HeaderMappingError:
                findings["sheets"][path.name].append(f"{sheet_name} (no header)")
                continue
            findings["sheets"][path.name].append(f"{sheet_name} (data)")
            findings["headers"][path.name] = [str(h) for h in header_row if h is not None]

            index = {
                field: header_row.index(column)
                for field, column in header_map.by_field.items()
            }

            def cell(row, field):
                position = index.get(field)
                if position is None or position >= len(row):
                    return None
                return row[position]

            for row in rows:
                if row is None or not any(v not in (None, "") for v in row):
                    continue
                for field in ("count_recorded", "count_closed"):
                    value = cell(row, field)
                    if value is None or isinstance(value, (int, float)):
                        continue
                    findings["non_numeric_counts"][str(value)] += 1
                findings["expired_flags"][str(cell(row, "offence_code_expired"))] += 1
                findings["outcome_types"][str(cell(row, "outcome_type"))] += 1
                findings["financial_years"][str(cell(row, "financial_year"))] += 1
                findings["quarters"][str(cell(row, "financial_quarter"))] += 1
                name = cell(row, "force_name")
                if canonical_force(name) is None:
                    findings["unmatched_forces"][text_of(name)] += 1

    print("\n" + "=" * 70)
    print("SURVEY")
    print("=" * 70)

    print("\nSheets per file:")
    for name, sheets in sorted(findings["sheets"].items()):
        print(f"  {name}")
        for sheet in sheets:
            print(f"      {sheet}")

    print("\nNon-numeric values in the two count columns:")
    if not findings["non_numeric_counts"]:
        print("  none")
    for value, count in findings["non_numeric_counts"].most_common():
        try:
            reading = to_int(value)
            verdict = f"reads as {reading}"
        except ValueError:
            verdict = "NOT HANDLED, would stop the build"
        print(f"  {value!r:40} {count:>9,}  {verdict}")

    print("\nOffence code expired flag values:")
    for value, count in findings["expired_flags"].most_common():
        print(f"  {value!r:40} {count:>9,}  {'expired' if truthy_flag(value) else 'current'}")

    print("\nOutcome type values:")
    unreadable = []
    for value, count in sorted(findings["outcome_types"].items()):
        try:
            int(value)
            note = ""
        except ValueError:
            note = "  NOT AN INTEGER, dropped as a subset row"
            unreadable.append(value)
        print(f"  {value!r:12} {count:>9,}{note}")

    print("\nForce names that do not match the canonical list:")
    if not findings["unmatched_forces"]:
        print("  none")
    for value, count in findings["unmatched_forces"].most_common(30):
        print(f"  {value!r:40} {count:>9,}")

    print("\nFinancial years:", ", ".join(sorted(findings["financial_years"])))
    print("Quarters:", ", ".join(sorted(findings["quarters"])))

    print("\nHeaders, where they differ between files:")
    seen = {}
    for name, headers in sorted(findings["headers"].items()):
        key = tuple(headers)
        seen.setdefault(key, []).append(name)
    for key, names in seen.items():
        print(f"  {len(names)} file(s), for example {names[0]}:")
        for header in key:
            print(f"      {header}")
    return findings


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--survey",
        action="store_true",
        help="report every distinct value the ETL has to interpret, across all "
        "files, and build nothing. Use this to find every surprise in one pass "
        "rather than one crash at a time.",
    )
    parser.add_argument(
        "--data-root",
        default=None,
        help="data directory to read and write, or 'fixture' for the "
        "synthetic tree. Defaults to data/.",
    )
    args = parser.parse_args()
    paths = resolve_data_root(args.data_root)
    args.raw_dir = paths.raw
    args.out_dir = paths.processed

    if not paths.manifest.exists():
        print(
            f"ERROR: {paths.manifest} is missing. Run etl/fetch.py first, so "
            "every derived figure is traceable to a recorded source file.",
            file=sys.stderr,
        )
        return 2

    manifest = read_json(paths.manifest)
    files = manifest.get("files", [])
    outcomes_files = [
        args.raw_dir / entry["filename"]
        for entry in files
        if entry.get("kind") == "outcomes_open_data"
    ]
    crime_files = [
        args.raw_dir / entry["filename"]
        for entry in files
        if entry.get("kind") == "police_force_area_crime"
    ]

    missing = [path for path in outcomes_files + crime_files if not path.exists()]
    if missing:
        print(
            f"ERROR: files named in the manifest are not in {paths.raw}: "
            + ", ".join(path.name for path in missing)
            + ". Run etl/fetch.py to rebuild them.",
            file=sys.stderr,
        )
        return 2
    if not outcomes_files:
        print("ERROR: the manifest lists no outcomes files.", file=sys.stderr)
        return 2

    log: list[str] = []

    if args.survey:
        survey(outcomes_files, log)
        return 0

    aggregator = Aggregator()
    for path in sorted(outcomes_files):
        process_outcomes_file(path, aggregator, log)

    if aggregator.unknown_forces:
        line = "Force names that did not match the canonical list:"
        log.append(line)
        print(line, flush=True)
        for name, count in sorted(aggregator.unknown_forces.items()):
            line = f"  {name!r}: {count} rows"
            log.append(line)
            print(line, flush=True)

    tables = build_tables(aggregator)
    denominators = process_force_area_crime(sorted(crime_files), log)

    years = sorted({row["fy"] for row in tables["force_year"]})
    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "provenance": manifest.get("provenance"),
        "manifest_generated_at": manifest.get("generated_at"),
        "attribution": manifest.get("attribution"),
        "default_basis": DEFAULT_BASIS,
        "count_bases": list(COUNT_BASES),
        "fraud_variants": list(FRAUD_VARIANTS),
        "count_columns": list(COUNT_COLUMNS),
        "oocd_outcome_types": list(OOCD_TYPES),
        "financial_years": years,
        "outcome_labels": {str(k): v for k, v in OUTCOME_LABELS.items()},
    }

    args.out_dir.mkdir(parents=True, exist_ok=True)
    for name, rows in tables.items():
        write_table(args.out_dir, name, rows, meta)
    write_json_compact(args.out_dir / "denominators.json", denominators)
    with (args.out_dir / "denominators.csv").open(
        "w", newline="", encoding="utf-8"
    ) as handle:
        writer = csv.DictWriter(handle, fieldnames=["force", "slug", "fy", "recorded_crime"])
        writer.writeheader()
        writer.writerows(denominators["rows"])

    coverage = {
        "rows_read": aggregator.rows_read,
        "rows_used": aggregator.rows_used,
        "negative_count_rows": aggregator.negative_counts,
        "negative_count_rows_excluded": aggregator.negative_counts_excluded,
        "negative_count_examples": aggregator.negative_examples,
        "negative_count_excluded_examples": aggregator.negative_excluded_examples,
        "negative_counts_by_force_year": {
            f"{force}|{financial_year}": count
            for (force, financial_year), count in sorted(aggregator.negative_by_key.items())
        },
        "subset_outcome_type_rows_dropped": dict(sorted(aggregator.subset_type_rows.items())),
        "duplicate_source_keys": aggregator.duplicate_source_keys,
        "duplicate_source_key_examples": aggregator.duplicate_examples,
        "data_not_provided": {
            f"{force}|{financial_year}": count
            for (force, financial_year), count in sorted(aggregator.not_provided.items())
        },
        "not_yet_assigned_rows_excluded": aggregator.not_assigned_rows,
        "not_yet_assigned_offences": dict(aggregator.not_assigned_counts),
        "central_fraud_body_rows_dropped": aggregator.central_fraud_rows,
        "expired_offence_code_rows": aggregator.expired_code_rows,
        "unknown_force_names": dict(sorted(aggregator.unknown_forces.items())),
        "outcome_type_years": {
            str(outcome_type): sorted(years)
            for outcome_type, years in sorted(aggregator.outcome_type_years.items())
        },
        "forces": sorted({row["force"] for row in tables["force_year"]}),
        "financial_years": years,
        "table_row_counts": {name: len(rows) for name, rows in tables.items()},
        "log": log,
    }
    write_json_compact(args.out_dir / "coverage.json", coverage)

    print("\nDerived tables written to", args.out_dir)
    for name, rows in tables.items():
        print(f"  {name}: {len(rows)} rows")
    print(f"  denominators: {len(denominators['rows'])} rows")
    print(f"Rows read {aggregator.rows_read}, rows used {aggregator.rows_used}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
